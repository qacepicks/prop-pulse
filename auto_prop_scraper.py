#!/usr/bin/env python3
# auto_prop_scraper.py — NBA Player Prop Auto Runner (PrizePicks Edition)
# ===============================================================
# Automatically fetches live NBA props from PrizePicks (main board only),
# runs PropPulse+ analysis, and exports results to Excel.
# ===============================================================

import os, time, json, requests, pandas as pd
from datetime import datetime
from prop_ev import analyze_single_prop, load_settings
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter


# ===============================================================
# 🧠 Auto-fetch PrizePicks Props (Main Board NBA)
# ===============================================================
def fetch_prizepicks_props():
    """Fetch live NBA props from PrizePicks API (main board only)"""
    # Use the proper PrizePicks API endpoint with filters for main board
    url = "https://api.prizepicks.com/projections"
    params = {
        "league_id": 7,          # NBA
        "per_page": 500,         # Get more props per request
        "single_stat": "true",   # Single stats only (no combos)
    }
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)",
        "Accept": "application/json, text/plain, */*",
        "Origin": "https://app.prizepicks.com",
        "Referer": "https://app.prizepicks.com/",
    }

    print("[PrizePicks] 🔄 Fetching live NBA projections...")
    try:
        r = requests.get(url, params=params, headers=headers, timeout=15)
        r.raise_for_status()
        data = r.json()
        total_props = len(data.get('data', []))
        print(f"[PrizePicks] ✅ Response OK — {total_props} props found.")
        
        # Debug: Show sample prop structure
        if total_props > 0:
            sample = data['data'][0]
            print(f"[DEBUG] Sample prop attributes: {list(sample.get('attributes', {}).keys())}")
    except Exception as e:
        print(f"[PrizePicks] ❌ Error fetching props: {e}")
        return []

    props = []
    included = {i["id"]: i for i in data.get("included", [])}
    
    # Debug counters
    filtered_counts = {
        "promo": 0,
        "flash_sale": 0,
        "special_desc": 0,
        "non_nba": 0,
        "bad_stat": 0,
        "bad_player": 0,
        "total_processed": 0
    }

    for proj in data.get("data", []):
        filtered_counts["total_processed"] += 1
        attrs = proj.get("attributes", {})
        rel = proj.get("relationships", {})
        player_rel = rel.get("new_player", {}).get("data", {})
        player_id = player_rel.get("id")

        # ⚠️ CRITICAL: Filter out non-main boards
        is_promo = attrs.get("is_promo", False)
        flash_sale = attrs.get("flash_sale_line_score")
        description = attrs.get("description", "").upper()
        odds_type = attrs.get("odds_type", "").upper()
        
        # Skip promotional/flash sale lines
        if is_promo:
            filtered_counts["promo"] += 1
            continue
        if flash_sale:
            filtered_counts["flash_sale"] += 1
            continue
        
        # Skip boards with special modifiers in description
        skip_keywords = ["GOBLIN", "DEMON", "FLASH", "SPECIAL", "BOOST"]
        if any(kw in description for kw in skip_keywords):
            filtered_counts["special_desc"] += 1
            continue
        
        # Skip non-standard odds types (e.g., boosted odds)
        if odds_type and odds_type not in ["", "STANDARD", "NORMAL"]:
            continue

        player_data = included.get(player_id, {}).get("attributes", {})
        player_name = player_data.get("display_name", "Unknown")
        stat_type = attrs.get("stat_type", "").upper()
        line_score = attrs.get("line_score", 0)

        # League filtering
        league_name = (
            attrs.get("league", "")
            or attrs.get("league_name", "")
            or included.get(player_id, {}).get("attributes", {}).get("league", "")
        ).upper()
        if "NBA" not in league_name and attrs.get("league_id") != 7:
            filtered_counts["non_nba"] += 1
            continue

        # Stat type mapping
        stat_map = {
    "POINTS": "PTS",
    "REBOUNDS": "REB",
    "ASSISTS": "AST",

    # Combined props
    "PTS+REB+AST": "PRA",
    "PTS+REB": "PR",          # NEW
    "PTS+AST": "PA",          # NEW
    "REB+AST": "REB+AST",

    # Threes
    "THREES": "FG3M",
    "3-PT MADE": "FG3M",
}

        stat = stat_map.get(stat_type, stat_type)

        # Only allow main stat types
        allowed = {"PTS", "REB", "AST", "PRA", "PR", "PA", "REB+AST", "FG3M"}
        if stat not in allowed:
            continue

        # Filter out unwanted prop types
        bad_keywords = [
            "COMBO", "FANTASY", "TURNOVER", "1ST", "HALF", "BLOCK", "STEAL",
            "SCORE", "TO", "DEF", "BLK", "STL"
        ]
        if any(b in stat_type for b in bad_keywords):
            continue

        # Hard filters for invalid players
        bad_names = ["A.J. Green", "Two-Way", "Unsigned", "Inactive", "Exhibit 10", "G League", "TBD", "Test"]
        if (
            line_score == 0
            or not player_name
            or player_name.strip() == ""
            or any(bad.lower() in player_name.lower() for bad in bad_names)
        ):
            filtered_counts["bad_player"] += 1
            continue

        props.append({
            "player": player_name,
            "stat": stat,
            "line": float(line_score),
            "odds": -110
        })

    # Print filter statistics
    print(f"[Filter Stats] Processed: {filtered_counts['total_processed']}")
    print(f"  ├─ Promo: {filtered_counts['promo']}")
    print(f"  ├─ Flash Sale: {filtered_counts['flash_sale']}")
    print(f"  ├─ Special Desc: {filtered_counts['special_desc']}")
    print(f"  ├─ Non-NBA: {filtered_counts['non_nba']}")
    print(f"  ├─ Bad Stat: {filtered_counts['bad_stat']}")
    print(f"  ├─ Bad Player: {filtered_counts['bad_player']}")
    print(f"  └─ Kept: {len(props)}")
    print(f"[PrizePicks] 🔍 Initial filter: {len(props)} valid NBA props")

    # ===============================================================
    # 🎯 Smart Deduplication (Keep Main Board Lines Only)
    # ===============================================================
    if not props:
        return []

    df = pd.DataFrame(props)
    
    # Additional filtering: Remove extreme outliers that suggest alt boards
    def is_likely_main_board(row):
        stat, line = row["stat"], row["line"]
        
        # Extreme lines are usually alternates
        outlier_thresholds = {
            "PTS": (line < 3 or line > 50),
            "REB": (line < 0.5 or line > 25),
            "AST": (line < 0.5 or line > 18),
            "PRA": (line < 8 or line > 65),
            "REB+AST": (line < 1 or line > 30),
            "FG3M": (line < 0.5 or line > 8)
        }
        
        if stat in outlier_thresholds and outlier_thresholds[stat]:
            return False
        
        # Half-point lines in assists/rebounds under 2 are usually gimmicks
        if stat in ["AST", "REB", "FG3M"] and line < 2.0 and (line % 1 == 0.5):
            return False
            
        return True
    
    df = df[df.apply(is_likely_main_board, axis=1)]
    print(f"[Filter] 🚫 Removed extreme outliers: {len(df)} props remain")
    
    # Define realistic line ranges for each stat type (main board thresholds)
    line_ranges = {
    "PTS": (5, 45),
    "REB": (1, 20),
    "AST": (1, 15),
    "PRA": (10, 60),
    "PR": (8, 50),        # NEW
    "PA": (5, 40),        # NEW
    "REB+AST": (2, 25),
    "FG3M": (0.5, 7)
}

    
    # Filter to realistic ranges first
    def is_main_board_line(row):
        stat = row["stat"]
        line = row["line"]
        if stat in line_ranges:
            min_val, max_val = line_ranges[stat]
            return min_val <= line <= max_val
        return True
    
    df = df[df.apply(is_main_board_line, axis=1)]
    print(f"[Filter] 📏 Filtered to realistic line ranges: {len(df)} props")
    
    # Group by player+stat and pick the most common line (mode)
    def select_main_line(group):
        if len(group) == 1:
            return group.iloc[0]
        
        # Get the most frequently occurring line (mode)
        line_counts = group["line"].value_counts()
        most_common_line = line_counts.index[0]
        max_count = line_counts.iloc[0]
        
        # Strong preference: line that appears multiple times
        if max_count >= 2:
            return group[group["line"] == most_common_line].iloc[0]
        
        # If all unique, look for "typical" main board values
        stat = group["stat"].iloc[0]
        lines = sorted(group["line"].values)
        
        # Main board lines cluster around typical values
        # Pick line closest to these common thresholds
        typical_values = {
            "PTS": [10.5, 12.5, 14.5, 16.5, 18.5, 20.5, 22.5, 25.5, 28.5, 30.5],
            "REB": [3.5, 4.5, 5.5, 6.5, 7.5, 8.5, 9.5, 10.5],
            "AST": [2.5, 3.5, 4.5, 5.5, 6.5, 7.5, 8.5],
            "PRA": [18.5, 20.5, 22.5, 25.5, 28.5, 30.5, 35.5],
            "REB+AST": [4.5, 5.5, 6.5, 7.5, 8.5, 9.5, 10.5],
            "FG3M": [1.5, 2.5, 3.5, 4.5]
        }

        if stat in typical_values:
            targets = typical_values[stat]
            min_dist = float("inf")
            best_line = group["line"].iloc[0]

            for line in group["line"]:
                dist = min(abs(line - t) for t in targets)
                if dist < min_dist:
                    min_dist = dist
                    best_line = line

            return group[group["line"] == best_line].iloc[0]

        # Fallback: pick median
        median_line = group["line"].median()
        closest_idx = (group["line"] - median_line).abs().idxmin()
        return group.loc[closest_idx]

    # 🔒 Safety check before deduplication
    if "stat" not in df.columns or "player" not in df.columns:
        print("[ERROR] Missing expected columns in DataFrame. Columns found:", df.columns.tolist())
        return []

    # Apply deduplication
    result_df = (
        df.groupby(["player", "stat"], as_index=False, group_keys=False)
        .apply(select_main_line)
        .reset_index(drop=True)
    )

    props = result_df.to_dict("records")
    print(f"[Filter] 🧩 Deduplicated alternate lines → {len(props)} main board props remain")

    # Debug output: Show sample of what we kept
    if len(props) > 0:
        print("\n[Sample] First 5 props after filtering:")
        for i, p in enumerate(props[:5], 1):
            print(f"  {i}. {p['player']:20s} {p['stat']:6s} {p['line']}")

    # Final sanity check
    props = [p for p in props if p["line"] > 0 and len(p["player"]) > 2]
    print(f"[PrizePicks] 🎯 Final count: {len(props)} NBA main board props\n")

    return props


# ===============================================================
# 📊 Excel Export with Dynamic Color Coding
# ===============================================================
def save_results_excel(df: pd.DataFrame, xlsx_path: str):
    """Write DataFrame to Excel with color-coded projections"""
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="NBA_PROPS")

    wb = load_workbook(xlsx_path)
    ws = wb["NBA_PROPS"]

    proj_col, line_col = None, None
    for c in range(1, ws.max_column + 1):
        val = str(ws.cell(row=1, column=c).value).lower()
        if "projection" in val:
            proj_col = c
        elif val == "line":
            line_col = c

    if proj_col and line_col:
        for row in range(2, ws.max_row + 1):
            proj = ws.cell(row=row, column=proj_col).value
            line = ws.cell(row=row, column=line_col).value
            if proj is None or line is None:
                continue

            neutral_zone = max(0.2, line * 0.05)
            diff = proj - line
            cell = ws.cell(row=row, column=proj_col)

            if abs(diff) <= neutral_zone:
                cell.fill = PatternFill(start_color="FFEB9C", fill_type="solid")  # yellow
            elif diff > neutral_zone:
                cell.fill = PatternFill(start_color="C6EFCE", fill_type="solid")  # green
            else:
                cell.fill = PatternFill(start_color="FFC7CE", fill_type="solid")  # red

    wb.save(xlsx_path)
    print(f"✅ Excel exported → {xlsx_path}")


# ===============================================================
# 🧠 Main Auto Runner
# ===============================================================
def main():
    print("=" * 60)
    print("🏀 PropPulse+ NBA Auto Analyzer (Auto Fetch Edition)")
    print("=" * 60)

    settings = load_settings()
    props = fetch_prizepicks_props()
    if not props:
        print("⚠️ No props found — try again later.")
        return

    results = []
    total = len(props)
    print(f"\n📊 Analyzing {total} props...\n")

    for i, prop in enumerate(props, start=1):
        player, stat, line, odds = prop["player"], prop["stat"], prop["line"], prop["odds"]
        print(f"[{i}/{total}] {player} — {stat} {line}")

        try:
            res = analyze_single_prop(player, stat, line, odds, settings, debug_mode=False)
            if res:
                proj = res.get("projection", 0)
                if proj > line * 2 or proj < line / 2:
                    proj = (proj + line) / 2  # clamp unrealistic

                results.append({
                    "Player": player,
                    "Stat": stat,
                    "Line": line,
                    "Odds": odds,
                    "Projection": proj,
                    "Model Prob": f"{res.get('p_model', 0)*100:.1f}%",
                    "Book Prob": f"{res.get('p_book', 0)*100:.1f}%",
                    "EV¢": f"{res.get('ev', 0)*100:+.1f}",
                    "Confidence": f"{res.get('confidence', 0):.2f}",
                    "Grade": res.get("grade", "N/A"),
                    "Result": res.get("result", "⚠️"),
                    "Opponent": res.get("opponent", "N/A"),
                    "Direction": res.get("direction", "N/A"),
                    "DvP Mult": f"{res.get('dvp_mult', 1.0):.3f}",
                    "Games": res.get("n_games", 0)
                })
        except Exception as e:
            print(f" ❌ Error analyzing {player}: {e}")

        time.sleep(0.2)

    if not results:
        print("⚠️ No successful analyses — check prop data.")
        return

    df = pd.DataFrame(results)
    df["EV_numeric"] = df["EV¢"].astype(str).str.replace("+", "").astype(float)

    # ===============================================================
    # 🧹 Filter out weak edges
    # ===============================================================
    df["Delta"] = abs(df["Projection"] - df["Line"])

    def is_strong(row):
        if row["Stat"] in ["PTS", "PRA"]:
            return row["Delta"] >= 1.0
        elif row["Stat"] in ["REB", "AST", "REB+AST"]:
            return row["Delta"] >= 0.8
        elif row["Stat"] in ["FG3M"]:
            return row["Delta"] >= 0.5
        return row["Delta"] >= 1.0

    df = df[df.apply(is_strong, axis=1)]
    print(f"[Filter] ⚖️ Removed weak edges — {len(df)} strong props remain")

    # ===============================================================
    # 🧠 Sort, trim, and export results
    # ===============================================================
    df = df.sort_values("EV_numeric", ascending=False).drop(columns=["EV_numeric"])
    df = df.head(220)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    xlsx_path = f"proppulse_results_{timestamp}.xlsx"
    save_results_excel(df, xlsx_path)

    print("\n🔥 Top EV Props:")
    print("=" * 60)
    for _, row in df.head(10).iterrows():
        print(f"{row['Player']:20s} {row['Stat']:6s} {row['Line']:5.1f} | "
              f"EV: {row['EV¢']:>6s}¢ | Grade: {row['Grade']}")
    print("=" * 60)
    print("\n✅ Analysis complete!\n")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n⚠️ Interrupted by user.")
    except Exception as e:
        import traceback
        print(f"❌ Fatal error: {e}")
        traceback.print_exc()